# fileName: excel_export.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def export_gruppe_excel(data, gruppe_name=None, veranstaltung_titel=None):
    """Exportiere Gruppendaten als formatiertes Excel (erfüllt alle Export-Anforderungen)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Teilnehmer"

    # Header-Informationen
    start_row = 1
    if veranstaltung_titel or gruppe_name:
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        v_titel = veranstaltung_titel if veranstaltung_titel else 'N/A'
        g_name = gruppe_name if gruppe_name else 'N/A'
        title_cell.value = f"Veranstaltung: {v_titel} - Gruppe: {g_name}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:G2')
        date_cell = ws['A2']
        date_cell.value = f"Exportiert am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal='center')
        
        start_row = 4

    # Spaltenüberschriften (entspricht den Anforderungen: Name, Matrikelnummer, Studiengang, Semester, E-Mail)
    headers = [
        "Name", "Matrikelnummer", "Studiengang",
        "Semester", "E-Mail", "Status", "Anmeldung am"
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Daten einfügen und Status farbig markieren
    status_colors = {
        "angemeldet": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "warteliste": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    }

    for row_num, row_data in enumerate(data, start_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            
            # Status-Spalte (Index 6) farbig markieren
            if col_num == 6 and value in status_colors:
                cell.fill = status_colors[value]

    # Spaltenbreiten anpassen
    column_widths = [25, 15, 20, 10, 30, 12, 18]
    for col_num, width in enumerate(column_widths, 1):
        # chr(64 + col_num) konvertiert 1->A, 2->B usw.
        ws.column_dimensions[chr(64 + col_num)].width = width 

    # Temporäre Datei speichern
    filename = f"temp_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.getpid()}.xlsx"
    wb.save(filename)
    return filename